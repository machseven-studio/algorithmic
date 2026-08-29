import io
import os
import tempfile
from flask import Flask, request, render_template_string, redirect, url_for, send_file
import openpyxl
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# --- HTML & CSS TEMPLATE ---
# Using Tailwind CSS via CDN for sleek, minimalistic, dark design
# No external CSS files needed; everything is in one file.

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Primitive</title>
    <!-- Tailwind CSS for sleek design -->
    <script src="https://cdn.tailwindcss.com"></script>
    <!-- Google Fonts: Inter for a modern, authoritative look -->
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;900&display=swap" rel="stylesheet">
    <style>
        body {
            font-family: 'Inter', sans-serif;
            background-color: #0f172a; /* Slate 900 */
            color: #f8fafc; /* Slate 50 */
        }
        .glass-panel {
            background: rgba(30, 41, 59, 0.7); /* Slate 800 with opacity */
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255, 255, 255, 0.1);
        }
        .input-field {
            background-color: #1e293b; /* Slate 800 */
            border: 1px solid #334155; /* Slate 700 */
            color: white;
            transition: all 0.3s ease;
        }
        .input-field:focus {
            border-color: #6366f1; /* Indigo 500 */
            outline: none;
            box-shadow: 0 0 0 2px rgba(99, 102, 241, 0.2);
        }
        .btn-primary {
            background-color: #4f46e5; /* Indigo 600 */
            transition: all 0.3s ease;
        }
        .btn-primary:hover {
            background-color: #4338ca; /* Indigo 700 */
            transform: translateY(-2px);
        }
        .label-text {
            color: #94a3b8; /* Slate 400 */
            font-size: 0.875rem;
            margin-bottom: 0.25rem;
        }
    </style>
</head>
<body class="min-h-screen flex flex-col items-center justify-center p-4">

    <!-- Header -->
    <header class="w-full max-w-4xl mb-8 flex justify-between items-center">
        <h1 class="text-4xl font-black tracking-tighter text-white">PRIMITIVE.</h1>
        <div class="text-sm font-medium text-slate-400">Automated Institute Management</div>
    </header>

    <!-- Main Content -->
    <main class="w-full max-w-4xl">
        
        <!-- Upload Section -->
        <div class="glass-panel rounded-2xl p-8 shadow-2xl">
            <div class="text-center mb-8">
                <h2 class="text-2xl font-bold mb-2">Upload Your Data</h2>
                <p class="text-slate-400 max-w-lg mx-auto">
                    Drop your Excel file below. Ensure your columns match the requirements below for optimal processing.
                </p>
            </div>

            <!-- Requirements Card -->
            <div class="bg-slate-800/50 rounded-xl p-6 mb-8 border border-slate-700/50">
                <h3 class="text-lg font-semibold text-indigo-400 mb-4 flex items-center">
                    <svg xmlns="http://www.w3.org/2000/svg" class="h-5 w-5 mr-2" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M13 16h-1v-4h-1m1-4h.01M21 12a9 9 0 11-18 0 9 9 0 0118 0z" />
                    </svg>
                    Required Excel Columns
                </h3>
                <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
                    <div>
                        <h4 class="font-bold text-slate-200 mb-2">For Seating Arrangement:</h4>
                        <ul class="text-sm text-slate-400 list-disc list-inside space-y-1">
                            <li><strong>Student Name</strong></li>
                            <li><strong>Batch</strong> (e.g., "CSE-2024")</li>
                            <li><strong>Roll Number</strong></li>
                            <li><strong>Room Preference</strong> (Optional)</li>
                        </ul>
                    </div>
                    <div>
                        <h4 class="font-bold text-slate-200 mb-2">For Timetable Generation:</h4>
                        <ul class="text-sm text-slate-400 list-disc list-inside space-y-1">
                            <li><strong>Teacher Name</strong></li>
                            <li><strong>Subject</strong></li>
                            <li><strong>Available Days</strong> (e.g., "Mon,Wed,Fri")</li>
                            <li><strong>Available Slots</strong> (e.g., "09:00-10:30")</li>
                            <li><strong>Batch</strong></li>
                        </ul>
                    </div>
                </div>
            </div>

            <!-- Upload Form -->
            <form action="/upload" method="post" enctype="multipart/form-data" class="flex flex-col items-center">
                <div class="w-full mb-6">
                    <label for="file-upload" class="flex flex-col items-center justify-center w-full h-32 border-2 border-dashed border-slate-600 rounded-xl cursor-pointer hover:border-indigo-500 hover:bg-slate-800/30 transition-all">
                        <div class="flex flex-col items-center justify-center pt-5 pb-6">
                            <svg class="w-10 h-10 mb-3 text-slate-400" fill="none" stroke="currentColor" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M7 16a4 4 0 01-.88-7.903A5 5 0 1115.9 6L16 6a5 5 0 011 9.9M15 13l-3-3m0 0l-3 3m3-3v12"></path></svg>
                            <p class="mb-2 text-sm text-slate-400"><span class="font-semibold">Click to upload</span> or drag and drop</p>
                            <p class="text-xs text-slate-500">XLSX, XLS (MAX. 10MB)</p>
                        </div>
                        <input id="file-upload" type="file" name="file" class="hidden" accept=".xlsx, .xls" required />
                    </label>
                </div>
                
                <button type="submit" class="btn-primary w-full py-4 px-6 rounded-xl text-white font-bold text-lg shadow-lg shadow-indigo-500/20">
                    PROCESS DATA
                </button>
            </form>
        </div>

        <!-- Loading/Processing State (Hidden by default) -->
        <div id="processing" class="hidden glass-panel rounded-2xl p-8 mt-6 text-center">
            <div class="animate-spin rounded-full h-12 w-12 border-b-2 border-indigo-500 mx-auto mb-4"></div>
            <h3 class="text-xl font-bold">Processing Algorithms...</h3>
            <p class="text-slate-400 mt-2">Optimizing seating and timetables.</p>
        </div>

        {% if result_html %}
        <div class="glass-panel rounded-2xl p-8 mt-6 shadow-2xl">
            <div class="flex justify-between items-center mb-6">
                <h2 class="text-2xl font-bold">Generated Output</h2>
                <a href="/" class="text-sm text-indigo-400 hover:text-indigo-300 underline">Upload New File</a>
            </div>
            
            <!-- Tabs for Seating and Timetable -->
            <div class="flex space-x-4 mb-6 border-b border-slate-700 pb-2">
                <button id="tab-seating" class="tab-btn text-indigo-400 border-b-2 border-indigo-500 font-semibold pb-2 px-4" onclick="showTab('seating')">Seating Arrangement</button>
                <button id="tab-timetable" class="tab-btn text-slate-400 hover:text-slate-200 pb-2 px-4" onclick="showTab('timetable')">Timetable</button>
            </div>

            <!-- Seating Content -->
            <div id="content-seating" class="tab-content">
                {{ result_html|safe }}
            </div>

            <!-- Timetable Content -->
            <div id="content-timetable" class="tab-content hidden">
                {{ timetable_html|default('<p class="text-slate-400">Timetable data not processed or available.</p>')|safe }}
            </div>
            
            <div class="mt-6 flex justify-end">
                <a href="/download" class="btn-primary px-6 py-2 rounded-lg text-white font-medium inline-flex items-center">
                    <svg xmlns="http://www.w3.org/2000/svg" class="h-5 w-5 mr-2" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4" />
                    </svg>
                    Download Report
                </a>
            </div>
        </div>
        {% endif %}

    </main>

    <footer class="w-full max-w-4xl mt-12 text-center text-slate-600 text-sm pb-8">
        &copy; 2026 Primitive. Automated Institute Management.
    </footer>

    <script>
        function showTab(tabName) {
            // Hide all contents
            document.getElementById('content-seating').classList.add('hidden');
            document.getElementById('content-timetable').classList.add('hidden');
            
            // Reset all tab styles
            document.getElementById('tab-seating').classList.remove('text-indigo-400', 'border-b-2', 'border-indigo-500', 'font-semibold');
            document.getElementById('tab-seating').classList.add('text-slate-400');
            document.getElementById('tab-timetable').classList.remove('text-indigo-400', 'border-b-2', 'border-indigo-500', 'font-semibold');
            document.getElementById('tab-timetable').classList.add('text-slate-400');

            // Show selected content and highlight tab
            if (tabName === 'seating') {
                document.getElementById('content-seating').classList.remove('hidden');
                document.getElementById('tab-seating').classList.add('text-indigo-400', 'border-b-2', 'border-indigo-500', 'font-semibold');
                document.getElementById('tab-seating').classList.remove('text-slate-400');
            } else {
                document.getElementById('content-timetable').classList.remove('hidden');
                document.getElementById('tab-timetable').classList.add('text-indigo-400', 'border-b-2', 'border-indigo-500', 'font-semibold');
                document.getElementById('tab-timetable').classList.remove('text-slate-400');
            }
        }
    </script>
</body>
</html>
"""

# --- LOGIC ---

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE, result_html=None, timetable_html=None)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)
    
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)
    
    if file:
        # Save file temporarily
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        file.save(temp_file.name)
        
        # Process Excel
        try:
            wb = openpyxl.load_workbook(temp_file.name)
            ws = wb.active
            
            # --- PARSE SEATING DATA ---
            seating_data = []
            headers = [cell.value for cell in ws[1]]
            
            # Find column indices
            col_student = next((i for i, h in enumerate(headers) if h and 'name' in h.lower()), None)
            col_batch = next((i for i, h in enumerate(headers) if h and 'batch' in h.lower()), None)
            col_roll = next((i for i, h in enumerate(headers) if h and 'roll' in h.lower()), None)
            
            if col_student is None or col_batch is None:
                raise ValueError("Excel must contain 'Student Name' and 'Batch' columns.")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[col_student]:
                    student = {
                        'name': row[col_student],
                        'batch': row[col_batch] if col_batch is not None else 'Unknown',
                        'roll': row[col_roll] if col_roll is not None else 'N/A'
                    }
                    seating_data.append(student)
            
            # --- MOCK TIMETABLE DATA ---
            # In a real app, you'd parse teacher/subject/day/slot columns here
            timetable_data = [
                {'Day': 'Monday', 'Time': '09:00 - 10:30', 'Batch': 'CSE-2024', 'Teacher': 'Dr. Smith', 'Subject': 'Math'},
                {'Day': 'Monday', 'Time': '10:30 - 12:00', 'Batch': 'ECE-2024', 'Teacher': 'Prof. Johnson', 'Subject': 'Physics'},
                {'Day': 'Tuesday', 'Time': '09:00 - 10:30', 'Batch': 'CSE-2024', 'Teacher': 'Dr. Smith', 'Subject': 'Math'},
                {'Day': 'Tuesday', 'Time': '10:30 - 12:00', 'Batch': 'ECE-2024', 'Teacher': 'Prof. Johnson', 'Subject': 'Physics'},
                {'Day': 'Wednesday', 'Time': '09:00 - 10:30', 'Batch': 'CSE-2024', 'Teacher': 'Dr. Smith', 'Subject': 'Math'},
                {'Day': 'Wednesday', 'Time': '10:30 - 12:00', 'Batch': 'ECE-2024', 'Teacher': 'Prof. Johnson', 'Subject': 'Physics'},
            ]

            # --- GENERATE HTML TABLES ---
            
            # Seating Table HTML
            seating_html = """
            <div class="overflow-x-auto">
                <table class="w-full text-left text-sm text-slate-300">
                    <thead class="bg-slate-800 text-xs uppercase text-slate-400">
                        <tr>
                            <th class="px-6 py-3">Roll No.</th>
                            <th class="px-6 py-3">Student Name</th>
                            <th class="px-6 py-3">Batch</th>
                            <th class="px-6 py-3">Seat No.</th>
                        </tr>
                    </thead>
                    <tbody class="divide-y divide-slate-700">
            """
            
            # Simple mock seating logic: Just list them out with a generated seat number
            for idx, student in enumerate(seating_data, 1):
                seating_html += f"""
                        <tr class="hover:bg-slate-700/50">
                            <td class="px-6 py-4 font-medium text-white">{student['roll']}</td>
                            <td class="px-6 py-4">{student['name']}</td>
                            <td class="px-6 py-4"><span class="bg-indigo-900 text-indigo-300 py-1 px-2 rounded text-xs">{student['batch']}</span></td>
                            <td class="px-6 py-4">{idx}</td>
                        </tr>
                """
            
            seating_html += """
                    </tbody>
                </table>
            </div>
            """

            # Timetable Table HTML
            timetable_html = """
            <div class="overflow-x-auto">
                <table class="w-full text-left text-sm text-slate-300">
                    <thead class="bg-slate-800 text-xs uppercase text-slate-400">
                        <tr>
                            <th class="px-6 py-3">Day</th>
                            <th class="px-6 py-3">Time</th>
                            <th class="px-6 py-3">Batch</th>
                            <th class="px-6 py-3">Teacher</th>
                            <th class="px-6 py-3">Subject</th>
                        </tr>
                    </thead>
                    <tbody class="divide-y divide-slate-700">
            """
            
            for entry in timetable_data:
                timetable_html += f"""
                        <tr class="hover:bg-slate-700/50">
                            <td class="px-6 py-4">{entry['Day']}</td>
                            <td class="px-6 py-4">{entry['Time']}</td>
                            <td class="px-6 py-4"><span class="bg-emerald-900 text-emerald-300 py-1 px-2 rounded text-xs">{entry['Batch']}</span></td>
                            <td class="px-6 py-4">{entry['Teacher']}</td>
                            <td class="px-6 py-4">{entry['Subject']}</td>
                        </tr>
                """
            
            timetable_html += """
                    </tbody>
                </table>
            </div>
            """

        except Exception as e:
            return render_template_string(HTML_TEMPLATE, 
                                          result_html=f'<div class="p-4 bg-red-900/30 border border-red-700 rounded-lg text-red-200">Error: {str(e)}</div>', 
                                          timetable_html='')
        finally:
            os.unlink(temp_file.name)

        return render_template_string(HTML_TEMPLATE, result_html=seating_html, timetable_html=timetable_html)

@app.route('/download')
def download():
    # Create a dummy Excel file for download
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seating"
    ws.append(["Roll No", "Name", "Batch", "Seat No"])
    # Add dummy data
    for i in range(1, 101):
        ws.append([f"R{i}", f"Student {i}", "BATCH-A", i])
    
    ws2 = wb.create_sheet(title="Timetable")
    ws2.append(["Day", "Time", "Batch", "Teacher", "Subject"])
    ws2.append(["Monday", "09:00-10:30", "CSE-2024", "Dr. Smith", "Math"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Primitive_Export.xlsx"
    )

if __name__ == '__main__':
    app.run(debug=True)
