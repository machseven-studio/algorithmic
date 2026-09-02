"""ALGORITHMIC institutional operations API.

The application deliberately uses PostgreSQL only.  Set DATABASE_URL (or
POSTGRES_URL) in the environment before starting the service, for example:

    postgresql://postgres:postgres@localhost:5432/algorithmic

The API is multi-tenant: every branch-owned record is checked against the
institute belonging to the authenticated session before it is read or
changed.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import secrets
from contextlib import contextmanager
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

import psycopg
from fastapi import Depends, FastAPI, HTTPException, Header, Request, UploadFile
from fastapi.responses import FileResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, EmailStr, Field
from psycopg.rows import dict_row
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
DATABASE_URL = os.getenv("DATABASE_URL") or os.getenv("POSTGRES_URL")
SESSION_LIFETIME_DAYS = 7
PBKDF2_ITERATIONS = 200_000
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
ALLOWED_UPLOAD_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx"}

MODULES = [
    "dashboard",
    "students",
    "teachers",
    "classrooms",
    "syllabus",
    "timetable",
    "attendance",
    "seating",
    "invigilation",
    "fees",
]
TABLES = {
    "students": "students",
    "teachers": "teachers",
    "classrooms": "classrooms",
    "syllabus": "syllabus",
    "attendance": "attendance",
    "invigilation": "invigilation",
    "fees": "fees",
}
RECORD_MODULES = list(TABLES)

# Designations are presets, not security boundaries. The selected permissions
# are stored with the user and are what the server enforces.
DESIGNATIONS = ["admin", "accountant", "teacher", "head", "staff", "read_only"]
DEFAULT_PERMISSIONS = {
    # Presets intentionally do not grant a non-boss every module. The boss
    # account is the only account with unrestricted operational access.
    "admin": {m: {"read": m not in {"fees", "seating"}, "write": m not in {"fees", "seating"}} for m in MODULES},
    "accountant": {
        m: {"read": m in {"fees", "students", "dashboard"}, "write": m == "fees"}
        for m in MODULES
    },
    "teacher": {
        m: {"read": m in {"students", "teachers", "syllabus", "timetable", "attendance", "dashboard"}, "write": m in {"attendance"}}
        for m in MODULES
    },
    "head": {m: {"read": m in {"dashboard", "students", "teachers", "classrooms", "syllabus", "timetable", "attendance", "seating", "invigilation"}, "write": m in {"students", "teachers", "classrooms", "syllabus", "timetable", "attendance", "seating", "invigilation"}} for m in MODULES},
    "staff": {m: {"read": m in {"students", "teachers", "classrooms", "syllabus", "attendance", "dashboard"}, "write": False} for m in MODULES},
    "read_only": {m: {"read": m in {"dashboard", "students"}, "write": False} for m in MODULES},
}

app = FastAPI(title="ALGORITHMIC", version="5.0.0")


# ---------------------------------------------------------------------------
# PostgreSQL
# ---------------------------------------------------------------------------


def _database_url() -> str:
    if not DATABASE_URL:
        raise RuntimeError(
            "DATABASE_URL is not configured. ALGORITHMIC now requires PostgreSQL; "
            "set DATABASE_URL before starting the API."
        )
    return DATABASE_URL


def get_conn():
    return psycopg.connect(_database_url(), row_factory=dict_row)


@contextmanager
def db_cursor():
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            yield conn, cur
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def init_db() -> None:
    """Create the PostgreSQL schema on startup.

    This is intentionally idempotent so a fresh Render/Fly/Railway database
    can boot without a separate migration command. PostgreSQL's IF NOT EXISTS
    keeps a deployment restart safe.
    """
    schema = """
    CREATE TABLE IF NOT EXISTS institutes (
        id BIGSERIAL PRIMARY KEY,
        institute_name TEXT NOT NULL,
        full_name TEXT NOT NULL,
        email TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        password_salt TEXT NOT NULL,
        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS staff_users (
        id BIGSERIAL PRIMARY KEY,
        institute_id BIGINT NOT NULL REFERENCES institutes(id) ON DELETE CASCADE,
        full_name TEXT NOT NULL,
        email TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        password_salt TEXT NOT NULL,
        designation TEXT NOT NULL,
        permissions JSONB NOT NULL DEFAULT '{}'::jsonb,
        permission TEXT NOT NULL DEFAULT 'read_only',
        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS sessions (
        token TEXT PRIMARY KEY,
        institute_id BIGINT NOT NULL REFERENCES institutes(id) ON DELETE CASCADE,
        staff_user_id BIGINT REFERENCES staff_users(id) ON DELETE CASCADE,
        expires_at TIMESTAMPTZ NOT NULL
    );
    CREATE TABLE IF NOT EXISTS branches (
        id BIGSERIAL PRIMARY KEY,
        institute_id BIGINT NOT NULL REFERENCES institutes(id) ON DELETE CASCADE,
        name TEXT NOT NULL,
        UNIQUE(institute_id, name)
    );
    CREATE TABLE IF NOT EXISTS students (
        id BIGSERIAL PRIMARY KEY,
        branch_id BIGINT NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
        name TEXT NOT NULL,
        email TEXT,
        batch TEXT NOT NULL,
        status TEXT DEFAULT 'Active',
        document TEXT,
        roll_number TEXT,
        parent_contact TEXT,
        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS teachers (
        id BIGSERIAL PRIMARY KEY,
        branch_id BIGINT NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
        name TEXT NOT NULL,
        subject TEXT,
        department TEXT,
        document TEXT,
        contact_number TEXT
    );
    CREATE TABLE IF NOT EXISTS classrooms (
        id BIGSERIAL PRIMARY KEY,
        branch_id BIGINT NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
        room_no TEXT NOT NULL,
        capacity INTEGER,
        building TEXT,
        document TEXT
    );
    CREATE TABLE IF NOT EXISTS syllabus (
        id BIGSERIAL PRIMARY KEY,
        branch_id BIGINT NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
        subject TEXT NOT NULL,
        semester TEXT,
        units INTEGER,
        document TEXT,
        topic TEXT,
        teacher_name TEXT,
        num_lectures INTEGER,
        lecture_date DATE,
        timing TEXT
    );
    CREATE TABLE IF NOT EXISTS attendance (
        id BIGSERIAL PRIMARY KEY,
        branch_id BIGINT NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
        student_id BIGINT REFERENCES students(id) ON DELETE CASCADE,
        student_name TEXT NOT NULL,
        date DATE NOT NULL,
        status TEXT NOT NULL CHECK (status IN ('Present', 'Absent')),
        lecture TEXT,
        document TEXT,
        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        UNIQUE(branch_id, student_id, date)
    );
    CREATE TABLE IF NOT EXISTS timetable_generations (
        id BIGSERIAL PRIMARY KEY,
        branch_id BIGINT NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
        batch_name TEXT NOT NULL,
        config JSONB NOT NULL,
        created_by BIGINT,
        created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS timetables_slots (
        id BIGSERIAL PRIMARY KEY,
        branch_id BIGINT NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
        generation_id BIGINT REFERENCES timetable_generations(id) ON DELETE CASCADE,
        batch_name TEXT NOT NULL,
        day TEXT NOT NULL,
        time_slot TEXT NOT NULL,
        lecture_number INTEGER NOT NULL,
        start_time TIME NOT NULL,
        end_time TIME NOT NULL,
        subject TEXT NOT NULL,
        teacher TEXT NOT NULL,
        room TEXT
    );
    CREATE TABLE IF NOT EXISTS invigilation (
        id BIGSERIAL PRIMARY KEY,
        branch_id BIGINT NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
        teacher_name TEXT NOT NULL,
        contact_number TEXT,
        exam_date DATE,
        room TEXT,
        document TEXT
    );
    CREATE TABLE IF NOT EXISTS fees (
        id BIGSERIAL PRIMARY KEY,
        branch_id BIGINT NOT NULL REFERENCES branches(id) ON DELETE CASCADE,
        student_id BIGINT REFERENCES students(id) ON DELETE SET NULL,
        student_name TEXT NOT NULL,
        amount_inr NUMERIC(12,2) NOT NULL DEFAULT 0,
        status TEXT NOT NULL DEFAULT 'Pending',
        due_date DATE,
        document TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_students_branch_batch ON students(branch_id, batch);
    CREATE INDEX IF NOT EXISTS idx_attendance_branch_date ON attendance(branch_id, date);
    CREATE INDEX IF NOT EXISTS idx_timetable_branch_batch ON timetables_slots(branch_id, batch_name);
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(schema)
        conn.commit()


@app.on_event("startup")
def startup() -> None:
    init_db()


# ---------------------------------------------------------------------------
# Common helpers and auth
# ---------------------------------------------------------------------------


def hash_password(password: str, salt: str) -> str:
    return hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), bytes.fromhex(salt), PBKDF2_ITERATIONS
    ).hex()


def create_session(institute_id: int, staff_user_id: Optional[int] = None) -> str:
    token = secrets.token_urlsafe(32)
    expires = datetime.now(timezone.utc) + timedelta(days=SESSION_LIFETIME_DAYS)
    with db_cursor() as (_, cur):
        cur.execute(
            "INSERT INTO sessions(token, institute_id, staff_user_id, expires_at) VALUES (%s, %s, %s, %s)",
            (token, institute_id, staff_user_id, expires),
        )
    return token


class CurrentInstitute(BaseModel):
    id: int
    institute_name: str
    full_name: str
    email: str
    is_owner: bool
    permission: str
    designation: str = "boss"
    permissions: dict[str, Any] = Field(default_factory=dict)


def normalise_permissions(designation: str, permissions: Optional[dict]) -> dict:
    base = json.loads(json.dumps(DEFAULT_PERMISSIONS.get(designation, DEFAULT_PERMISSIONS["staff"])))
    if permissions:
        for module in MODULES:
            if module in permissions and isinstance(permissions[module], dict):
                base[module]["read"] = bool(permissions[module].get("read", False))
                base[module]["write"] = bool(permissions[module].get("write", False))
    # Nobody except the institute owner can ever receive user-management access.
    base.pop("users", None)
    return base


def user_has_access(user: CurrentInstitute, module: str, write: bool = False) -> bool:
    if user.is_owner:
        return True
    entry = user.permissions.get(module, {})
    if write:
        return bool(entry.get("write", False))
    return bool(entry.get("read", False) or entry.get("write", False))


def require_module_access(user: CurrentInstitute, module: str, write: bool = False) -> None:
    if module not in MODULES:
        raise HTTPException(status_code=400, detail="Invalid module")
    if not user_has_access(user, module, write):
        action = "edit" if write else "view"
        raise HTTPException(status_code=403, detail=f"Your designation cannot {action} the {module} module")


def model_data(model: BaseModel, **kwargs) -> dict:
    """Support both Pydantic 1 and 2 during a rolling deployment."""
    if hasattr(model, "model_dump"):
        return model.model_dump(**kwargs)
    return model.dict(**kwargs)


def get_current_institute(authorization: Optional[str] = Header(None)) -> CurrentInstitute:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = authorization.split(" ", 1)[1]
    with db_cursor() as (_, cur):
        cur.execute(
            """SELECT s.*, i.institute_name, i.full_name AS owner_name, i.email AS owner_email
               FROM sessions s JOIN institutes i ON i.id = s.institute_id
               WHERE s.token = %s""",
            (token,),
        )
        session = cur.fetchone()
        if not session:
            raise HTTPException(status_code=401, detail="Invalid or expired session")
        expires = session["expires_at"]
        if expires.tzinfo is None:
            expires = expires.replace(tzinfo=timezone.utc)
        if expires < datetime.now(timezone.utc):
            cur.execute("DELETE FROM sessions WHERE token = %s", (token,))
            raise HTTPException(status_code=401, detail="Session expired, please log in again")
        if session["staff_user_id"]:
            cur.execute("SELECT * FROM staff_users WHERE id = %s", (session["staff_user_id"],))
            staff = cur.fetchone()
            if not staff:
                raise HTTPException(status_code=401, detail="Invalid session")
            designation = staff["designation"]
            return CurrentInstitute(
                id=session["institute_id"], institute_name=session["institute_name"],
                full_name=staff["full_name"], email=staff["email"], is_owner=False,
                permission=staff["permission"], designation=designation,
                permissions=normalise_permissions(designation, staff["permissions"]),
            )
        return CurrentInstitute(
            id=session["institute_id"], institute_name=session["institute_name"],
            full_name=session["owner_name"], email=session["owner_email"], is_owner=True,
            permission="owner", designation="boss",
            permissions={m: {"read": True, "write": True} for m in MODULES},
        )


def require_write_access(user: CurrentInstitute = Depends(get_current_institute)) -> CurrentInstitute:
    if not user.is_owner and user.permission == "read_only":
        raise HTTPException(status_code=403, detail="Your account has read-only access")
    return user


def require_owner(user: CurrentInstitute = Depends(get_current_institute)) -> CurrentInstitute:
    if not user.is_owner:
        raise HTTPException(status_code=403, detail="Only the boss can manage users")
    return user


def verify_branch_ownership(branch_id: int, institute_id: int) -> None:
    with db_cursor() as (_, cur):
        cur.execute("SELECT id FROM branches WHERE id = %s AND institute_id = %s", (branch_id, institute_id))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Branch not found")


def verify_record_ownership(module: str, record_id: int, institute_id: int) -> dict:
    table = TABLES.get(module)
    if not table:
        raise HTTPException(status_code=400, detail="This module does not contain editable records")
    with db_cursor() as (_, cur):
        cur.execute(
            f"""SELECT t.* FROM {table} t JOIN branches b ON b.id=t.branch_id
                WHERE t.id=%s AND b.institute_id=%s""", (record_id, institute_id)
        )
        row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Record not found")
    return row


def parse_iso_date(value: Any) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid date: {value}")


def parse_clock(value: str) -> time:
    raw = str(value or "").strip().upper().replace(".", "")
    raw = re.sub(r"\s+", " ", raw)
    for fmt in ("%H:%M", "%I:%M %p", "%I %p"):
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            pass
    raise HTTPException(status_code=400, detail=f"Invalid time '{value}'. Use HH:MM or HH:MM AM/PM.")


def timing_from_input(value: Any, index: int) -> dict:
    if isinstance(value, str):
        text = value.strip()
        parts = re.split(r"\s*(?:-|–|to)\s*", text, maxsplit=1, flags=re.I)
        if len(parts) != 2:
            raise HTTPException(status_code=400, detail=f"Timing {index + 1} must contain a start and end time")
        start_text, end_text = parts
        number = index + 1
    elif isinstance(value, dict):
        number = int(value.get("lecture_number", index + 1))
        start_text = value.get("start_time") or value.get("start")
        end_text = value.get("end_time") or value.get("end")
    else:
        raise HTTPException(status_code=400, detail="Each lecture timing must be text or an object")
    start = parse_clock(str(start_text))
    end = parse_clock(str(end_text))
    if start >= end:
        raise HTTPException(status_code=400, detail=f"Lecture {number} must end after it starts")
    return {
        "lecture_number": number,
        "start_time": start.isoformat(timespec="minutes"),
        "end_time": end.isoformat(timespec="minutes"),
        "label": f"{start.strftime('%I:%M %p')} – {end.strftime('%I:%M %p')}".replace(" 0", " "),
    }


def save_upload(file: Optional[UploadFile]) -> Optional[str]:
    if not file or not file.filename:
        return None
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported file type '{ext}'")
    content = file.file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail="File too large (maximum 5 MB)")
    name = f"{secrets.token_hex(16)}{ext}"
    (UPLOAD_DIR / name).write_bytes(content)
    return name


# ---------------------------------------------------------------------------
# Auth and institute profile
# ---------------------------------------------------------------------------

class SignupRequest(BaseModel):
    institute_name: str
    full_name: str
    email: EmailStr
    password: str


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


@app.post("/api/auth/signup")
def signup(req: SignupRequest):
    if len(req.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    salt = secrets.token_hex(16)
    with db_cursor() as (_, cur):
        try:
            cur.execute(
                "INSERT INTO institutes(institute_name,full_name,email,password_hash,password_salt) VALUES(%s,%s,%s,%s,%s) RETURNING id",
                (req.institute_name.strip(), req.full_name.strip(), req.email.lower(), hash_password(req.password, salt), salt),
            )
            institute_id = cur.fetchone()["id"]
            cur.execute("INSERT INTO branches(institute_id,name) VALUES(%s,%s)", (institute_id, "Main Campus"))
        except psycopg.errors.UniqueViolation:
            raise HTTPException(status_code=400, detail="An account with this email already exists")
    return {"token": create_session(institute_id), "institute_name": req.institute_name, "full_name": req.full_name, "is_owner": True, "permission": "owner", "designation": "boss"}


@app.post("/api/auth/login")
def login(req: LoginRequest):
    with db_cursor() as (_, cur):
        cur.execute("SELECT * FROM institutes WHERE email=%s", (req.email.lower(),))
        institute = cur.fetchone()
        if institute and secrets.compare_digest(hash_password(req.password, institute["password_salt"]), institute["password_hash"]):
            return {"token": create_session(institute["id"]), "institute_name": institute["institute_name"], "full_name": institute["full_name"], "is_owner": True, "permission": "owner", "designation": "boss"}
        cur.execute("SELECT * FROM staff_users WHERE email=%s", (req.email.lower(),))
        staff = cur.fetchone()
        if staff and secrets.compare_digest(hash_password(req.password, staff["password_salt"]), staff["password_hash"]):
            cur.execute("SELECT institute_name FROM institutes WHERE id=%s", (staff["institute_id"],))
            parent = cur.fetchone()
            return {"token": create_session(staff["institute_id"], staff["id"]), "institute_name": parent["institute_name"] if parent else "", "full_name": staff["full_name"], "is_owner": False, "permission": staff["permission"], "designation": staff["designation"], "permissions": normalise_permissions(staff["designation"], staff["permissions"])}
    raise HTTPException(status_code=401, detail="Invalid email or password")


@app.get("/api/auth/me")
def whoami(user: CurrentInstitute = Depends(get_current_institute)):
    return model_data(user)


@app.post("/api/auth/logout")
def logout(authorization: Optional[str] = Header(None)):
    if authorization and authorization.startswith("Bearer "):
        with db_cursor() as (_, cur):
            cur.execute("DELETE FROM sessions WHERE token=%s", (authorization.split(" ", 1)[1],))
    return {"status": "logged out"}


class InstituteNameUpdate(BaseModel):
    institute_name: str


@app.patch("/api/institute/name")
def update_institute_name(req: InstituteNameUpdate, user: CurrentInstitute = Depends(require_write_access)):
    name = req.institute_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Institute name cannot be empty")
    with db_cursor() as (_, cur):
        cur.execute("UPDATE institutes SET institute_name=%s WHERE id=%s", (name, user.id))
    return {"institute_name": name}


# ---------------------------------------------------------------------------
# Manage users / designations / privileges
# ---------------------------------------------------------------------------

class StaffUserCreate(BaseModel):
    full_name: str
    email: EmailStr
    password: str
    designation: str
    permissions: Optional[dict[str, Any]] = None


class StaffUserUpdate(BaseModel):
    full_name: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = None
    designation: Optional[str] = None
    permissions: Optional[dict[str, Any]] = None
    # Legacy compatibility: older clients sent only permission=edit/read_only.
    permission: Optional[str] = None


@app.get("/api/designations")
def get_designations(user: CurrentInstitute = Depends(require_owner)):
    return {"designations": DESIGNATIONS, "modules": MODULES, "presets": DEFAULT_PERMISSIONS}


@app.get("/api/users")
def list_users(user: CurrentInstitute = Depends(require_owner)):
    with db_cursor() as (_, cur):
        cur.execute("SELECT id,full_name,email,designation,permissions,permission,created_at,updated_at FROM staff_users WHERE institute_id=%s ORDER BY full_name", (user.id,))
        rows = cur.fetchall()
    return [dict(r, permissions=normalise_permissions(r["designation"], r["permissions"])) for r in rows]


def safe_designation(value: str) -> str:
    value = value.strip().lower()
    if value not in DESIGNATIONS:
        raise HTTPException(status_code=400, detail=f"Designation must be one of: {', '.join(DESIGNATIONS)}")
    return value


@app.post("/api/users")
def add_user(req: StaffUserCreate, user: CurrentInstitute = Depends(require_owner)):
    designation = safe_designation(req.designation)
    if len(req.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    permissions = normalise_permissions(designation, req.permissions)
    any_write = any(v.get("write") for v in permissions.values())
    salt = secrets.token_hex(16)
    with db_cursor() as (_, cur):
        try:
            cur.execute(
                """INSERT INTO staff_users(institute_id,full_name,email,password_hash,password_salt,designation,permissions,permission)
                   VALUES(%s,%s,%s,%s,%s,%s,%s::jsonb,%s) RETURNING id""",
                (user.id, req.full_name.strip(), req.email.lower(), hash_password(req.password, salt), salt, designation, json.dumps(permissions), "edit" if any_write else "read_only"),
            )
            new_id = cur.fetchone()["id"]
        except psycopg.errors.UniqueViolation:
            raise HTTPException(status_code=400, detail="A user with this email already exists")
    return {"id": new_id, "full_name": req.full_name, "email": req.email.lower(), "designation": designation, "permissions": permissions}


@app.patch("/api/users/{user_id}")
def update_user(user_id: int, req: StaffUserUpdate, user: CurrentInstitute = Depends(require_owner)):
    with db_cursor() as (_, cur):
        cur.execute("SELECT * FROM staff_users WHERE id=%s AND institute_id=%s", (user_id, user.id))
        old = cur.fetchone()
        if not old:
            raise HTTPException(status_code=404, detail="User not found")
        designation = safe_designation(req.designation) if req.designation is not None else old["designation"]
        permissions = normalise_permissions(designation, req.permissions if req.permissions is not None else old["permissions"])
        if req.permission is not None:
            if req.permission not in {"edit", "read_only"}:
                raise HTTPException(status_code=400, detail="permission must be edit or read_only")
            for module in MODULES:
                permissions[module]["write"] = req.permission == "edit" and permissions[module]["read"]
        fields, values = ["full_name=%s", "email=%s", "designation=%s", "permissions=%s::jsonb", "permission=%s", "updated_at=NOW()"], [req.full_name.strip() if req.full_name else old["full_name"], str(req.email).lower() if req.email else old["email"], designation, json.dumps(permissions), "edit" if any(v.get("write") for v in permissions.values()) else "read_only"]
        if req.password:
            if len(req.password) < 8:
                raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
            salt = secrets.token_hex(16)
            fields += ["password_hash=%s", "password_salt=%s"]
            values += [hash_password(req.password, salt), salt]
        values += [user_id, user.id]
        try:
            cur.execute(f"UPDATE staff_users SET {', '.join(fields)} WHERE id=%s AND institute_id=%s", values)
        except psycopg.errors.UniqueViolation:
            raise HTTPException(status_code=400, detail="A user with this email already exists")
        cur.execute("DELETE FROM sessions WHERE staff_user_id=%s", (user_id,))
    return {"id": user_id, "status": "updated"}


@app.delete("/api/users/{user_id}")
def remove_user(user_id: int, user: CurrentInstitute = Depends(require_owner)):
    with db_cursor() as (_, cur):
        cur.execute("DELETE FROM staff_users WHERE id=%s AND institute_id=%s", (user_id, user.id))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="User not found")
    return {"status": "removed"}


# ---------------------------------------------------------------------------
# Branches and record CRUD
# ---------------------------------------------------------------------------

class BranchCreate(BaseModel):
    name: str


class BranchUpdate(BaseModel):
    name: str


@app.get("/api/branches")
def get_branches(user: CurrentInstitute = Depends(get_current_institute)):
    with db_cursor() as (_, cur):
        cur.execute("SELECT id,name FROM branches WHERE institute_id=%s ORDER BY id", (user.id,))
        return cur.fetchall()


@app.post("/api/branches")
def add_branch(req: BranchCreate, user: CurrentInstitute = Depends(require_write_access)):
    name = req.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Branch name cannot be empty")
    try:
        with db_cursor() as (_, cur):
            cur.execute("INSERT INTO branches(institute_id,name) VALUES(%s,%s) RETURNING id,name", (user.id, name))
            return cur.fetchone()
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=400, detail="Branch already exists")


@app.patch("/api/branches/{branch_id}")
def edit_branch(branch_id: int, req: BranchUpdate, user: CurrentInstitute = Depends(require_write_access)):
    verify_branch_ownership(branch_id, user.id)
    try:
        with db_cursor() as (_, cur):
            cur.execute("UPDATE branches SET name=%s WHERE id=%s RETURNING id,name", (req.name.strip(), branch_id))
            return cur.fetchone()
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=400, detail="Branch already exists")


def clean_record_data(module: str, data: dict) -> dict:
    fields = {
        "students": ["name", "email", "batch", "status", "roll_number", "parent_contact", "document"],
        "teachers": ["name", "subject", "department", "contact_number", "document"],
        "classrooms": ["room_no", "capacity", "building", "document"],
        "syllabus": ["subject", "semester", "units", "topic", "teacher_name", "num_lectures", "lecture_date", "timing", "document"],
        "attendance": ["student_id", "student_name", "date", "status", "lecture", "document"],
        "invigilation": ["teacher_name", "contact_number", "exam_date", "room", "document"],
        "fees": ["student_id", "student_name", "amount_inr", "status", "due_date", "document"],
    }[module]
    cleaned = {k: data.get(k) for k in fields if k in data}
    for key in ("lecture_date", "exam_date", "due_date", "date"):
        if key in cleaned:
            cleaned[key] = parse_iso_date(cleaned[key])
    if "status" in cleaned and module == "attendance" and cleaned["status"] not in {"Present", "Absent"}:
        raise HTTPException(status_code=400, detail="Attendance status must be Present or Absent")
    return cleaned


def insert_record(module: str, branch_id: int, data: dict, document: Optional[str] = None) -> int:
    if module == "students":
        cols = ["branch_id", "name", "email", "batch", "status", "roll_number", "parent_contact", "document"]
    elif module == "teachers":
        cols = ["branch_id", "name", "subject", "department", "contact_number", "document"]
    elif module == "classrooms":
        cols = ["branch_id", "room_no", "capacity", "building", "document"]
    elif module == "syllabus":
        cols = ["branch_id", "subject", "semester", "units", "topic", "teacher_name", "num_lectures", "lecture_date", "timing", "document"]
    elif module == "attendance":
        cols = ["branch_id", "student_id", "student_name", "date", "status", "lecture", "document"]
    elif module == "invigilation":
        cols = ["branch_id", "teacher_name", "contact_number", "exam_date", "room", "document"]
    elif module == "fees":
        cols = ["branch_id", "student_id", "student_name", "amount_inr", "status", "due_date", "document"]
    else:
        raise HTTPException(status_code=400, detail="Invalid record module")
    data = clean_record_data(module, data)
    if module == "students":
        data.setdefault("status", "Active")
    if module == "fees":
        data.setdefault("status", "Pending")
    if document:
        data["document"] = document
    values = [branch_id] + [data.get(c) for c in cols[1:]]
    placeholders = ",".join(["%s"] * len(cols))
    with db_cursor() as (_, cur):
        cur.execute(f"INSERT INTO {TABLES[module]}({','.join(cols)}) VALUES({placeholders}) RETURNING id", values)
        return cur.fetchone()["id"]


@app.get("/api/records/{module}/{branch_id}")
def get_records(module: str, branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    if module not in RECORD_MODULES:
        raise HTTPException(status_code=400, detail="Invalid record module")
    require_module_access(user, module)
    verify_branch_ownership(branch_id, user.id)
    with db_cursor() as (_, cur):
        cur.execute(f"SELECT * FROM {TABLES[module]} WHERE branch_id=%s ORDER BY id DESC", (branch_id,))
        return cur.fetchall()


@app.post("/api/records/{module}")
async def add_record(
    module: str,
    request: Request,
    user: CurrentInstitute = Depends(require_write_access),
):
    if module not in RECORD_MODULES:
        raise HTTPException(status_code=400, detail="Invalid record module")
    require_module_access(user, module, write=True)
    content_type = request.headers.get("content-type", "")
    file = None
    if "multipart/form-data" in content_type:
        form = await request.form()
        branch_id = form.get("branch_id")
        data = json.loads(form.get("data_json", "{}"))
        file = form.get("file")
    else:
        payload = await request.json()
        branch_id = payload.pop("branch_id", None)
        data = payload
    if not branch_id:
        raise HTTPException(status_code=400, detail="branch_id is required")
    verify_branch_ownership(int(branch_id), user.id)
    document = save_upload(file) if file is not None and hasattr(file, "filename") else None
    try:
        record_id = insert_record(module, int(branch_id), data, document)
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=400, detail="That record conflicts with an existing record")
    return {"id": record_id, "status": "success"}


@app.patch("/api/records/{module}/{record_id}")
async def edit_record(module: str, record_id: int, request: Request, user: CurrentInstitute = Depends(require_write_access)):
    if module not in RECORD_MODULES:
        raise HTTPException(status_code=400, detail="Invalid record module")
    require_module_access(user, module, write=True)
    existing = verify_record_ownership(module, record_id, user.id)
    content_type = request.headers.get("content-type", "")
    file = None
    if "multipart/form-data" in content_type:
        form = await request.form()
        data = json.loads(form.get("data_json", "{}"))
        file = form.get("file")
    else:
        data = await request.json()
    cleaned = clean_record_data(module, data)
    if file is not None and hasattr(file, "filename"):
        cleaned["document"] = save_upload(file)
    if not cleaned:
        raise HTTPException(status_code=400, detail="No editable fields supplied")
    set_parts, values = [], []
    for field, value in cleaned.items():
        set_parts.append(f"{field}=%s")
        values.append(value)
    values.append(record_id)
    with db_cursor() as (_, cur):
        cur.execute(f"UPDATE {TABLES[module]} SET {', '.join(set_parts)} WHERE id=%s", values)
    return {"id": record_id, "status": "updated"}


@app.delete("/api/records/{module}/{record_id}")
def delete_record(module: str, record_id: int, user: CurrentInstitute = Depends(require_write_access)):
    if module not in RECORD_MODULES:
        raise HTTPException(status_code=400, detail="Invalid record module")
    require_module_access(user, module, write=True)
    verify_record_ownership(module, record_id, user.id)
    with db_cursor() as (_, cur):
        cur.execute(f"DELETE FROM {TABLES[module]} WHERE id=%s", (record_id,))
    return {"status": "deleted"}


BULK_COLUMNS = {
    "students": ["name", "email", "batch", "status", "roll_number", "parent_contact"],
    "teachers": ["name", "subject", "department", "contact_number"],
    "classrooms": ["room_no", "capacity", "building"],
    "syllabus": ["subject", "semester", "units", "topic", "teacher_name", "num_lectures", "lecture_date", "timing"],
    "attendance": ["student_id", "student_name", "date", "status", "lecture"],
    "invigilation": ["teacher_name", "contact_number", "exam_date", "room"],
    "fees": ["student_id", "student_name", "amount_inr", "status", "due_date"],
}


@app.post("/api/records/{module}/bulk")
async def bulk_import_records(module: str, request: Request, user: CurrentInstitute = Depends(require_write_access)):
    """Keep the original CSV import workflow while using PostgreSQL CRUD."""
    if module not in RECORD_MODULES:
        raise HTTPException(status_code=400, detail="Invalid record module")
    require_module_access(user, module, write=True)
    form = await request.form()
    branch_id = form.get("branch_id")
    file = form.get("file")
    if not branch_id or file is None or not hasattr(file, "filename"):
        raise HTTPException(status_code=400, detail="branch_id and a CSV file are required")
    verify_branch_ownership(int(branch_id), user.id)
    if not str(file.filename).lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Please upload a .csv file")
    raw = await file.read()
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail="File too large (maximum 5 MB)")
    try:
        reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig")))
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Please save the CSV as UTF-8")
    expected = BULK_COLUMNS[module]
    fieldnames = [str(field).strip() for field in (reader.fieldnames or [])]
    if not set(expected).issubset(set(fieldnames)):
        raise HTTPException(status_code=400, detail=f"CSV must include: {', '.join(expected)}")
    inserted = 0
    for raw_row in reader:
        row = {str(k).strip(): (v.strip() if isinstance(v, str) else v) for k, v in raw_row.items() if k is not None}
        if not any(row.get(k) for k in expected):
            continue
        insert_record(module, int(branch_id), row)
        inserted += 1
    if not inserted:
        raise HTTPException(status_code=400, detail="No valid rows found in that file")
    return {"status": "success", "inserted": inserted}


# Compatibility aliases used by earlier clients. New clients use /api/records.
async def alias_create(module: str, request: Request, user: CurrentInstitute):
    payload = await request.json()
    branch_id = payload.pop("branch_id", None)
    if not branch_id:
        raise HTTPException(status_code=400, detail="branch_id is required")
    require_module_access(user, module, write=True)
    verify_branch_ownership(int(branch_id), user.id)
    return {"id": insert_record(module, int(branch_id), payload), "status": "success"}


@app.post("/api/students")
async def create_student_alias(request: Request, user: CurrentInstitute = Depends(require_write_access)):
    return await alias_create("students", request, user)


@app.post("/api/teachers")
async def create_teacher_alias(request: Request, user: CurrentInstitute = Depends(require_write_access)):
    return await alias_create("teachers", request, user)


@app.post("/api/classrooms")
async def create_classroom_alias(request: Request, user: CurrentInstitute = Depends(require_write_access)):
    return await alias_create("classrooms", request, user)


@app.post("/api/syllabus")
async def create_syllabus_alias(request: Request, user: CurrentInstitute = Depends(require_write_access)):
    return await alias_create("syllabus", request, user)


@app.post("/api/invigilators")
async def create_invigilator_alias(request: Request, user: CurrentInstitute = Depends(require_write_access)):
    return await alias_create("invigilation", request, user)


@app.post("/api/fees")
async def create_fee_alias(request: Request, user: CurrentInstitute = Depends(require_write_access)):
    return await alias_create("fees", request, user)


@app.get("/api/students")
def get_students_alias(branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    return get_records("students", branch_id, user)


@app.get("/api/teachers")
def get_teachers_alias(branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    return get_records("teachers", branch_id, user)


@app.get("/api/classrooms")
def get_classrooms_alias(branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    return get_records("classrooms", branch_id, user)


@app.get("/api/syllabus")
def get_syllabus_alias(branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    return get_records("syllabus", branch_id, user)


# ---------------------------------------------------------------------------
# Attendance and reports
# ---------------------------------------------------------------------------

class AttendanceMarkRequest(BaseModel):
    branch_id: int
    student_id: Optional[int] = None
    student_name: Optional[str] = None
    date: str
    status: str
    lecture: Optional[str] = None


@app.get("/api/student-batches")
def student_batches(branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    require_module_access(user, "students")
    verify_branch_ownership(branch_id, user.id)
    with db_cursor() as (_, cur):
        cur.execute("SELECT DISTINCT batch FROM students WHERE branch_id=%s AND batch IS NOT NULL ORDER BY batch", (branch_id,))
        return [r["batch"] for r in cur.fetchall()]


@app.get("/api/attendance/students")
def attendance_students(branch_id: int, batch: str, date: str, q: str = "", user: CurrentInstitute = Depends(get_current_institute)):
    require_module_access(user, "attendance")
    verify_branch_ownership(branch_id, user.id)
    with db_cursor() as (_, cur):
        cur.execute(
            """SELECT s.id,s.name,s.email,s.batch,s.roll_number,s.parent_contact,
                      a.status AS attendance_status
               FROM students s LEFT JOIN attendance a ON a.student_id=s.id AND a.date=%s
               WHERE s.branch_id=%s AND s.batch=%s AND (%s='' OR LOWER(s.name) LIKE LOWER(%s))
               ORDER BY s.name""",
            (parse_iso_date(date), branch_id, batch, q, f"%{q}%"),
        )
        return cur.fetchall()


@app.post("/api/attendance/mark")
def mark_attendance(req: AttendanceMarkRequest, user: CurrentInstitute = Depends(require_write_access)):
    require_module_access(user, "attendance", write=True)
    verify_branch_ownership(req.branch_id, user.id)
    if req.status not in {"Present", "Absent"}:
        raise HTTPException(status_code=400, detail="Status must be Present or Absent")
    with db_cursor() as (_, cur):
        student = None
        if req.student_id:
            cur.execute("SELECT id,name,batch FROM students WHERE id=%s AND branch_id=%s", (req.student_id, req.branch_id))
            student = cur.fetchone()
        elif req.student_name:
            cur.execute("SELECT id,name,batch FROM students WHERE name=%s AND branch_id=%s LIMIT 1", (req.student_name, req.branch_id))
            student = cur.fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="Student not found in this branch")
        cur.execute(
            """INSERT INTO attendance(branch_id,student_id,student_name,date,status,lecture)
               VALUES(%s,%s,%s,%s,%s,%s)
               ON CONFLICT(branch_id,student_id,date) DO UPDATE SET student_name=EXCLUDED.student_name,status=EXCLUDED.status,lecture=EXCLUDED.lecture""",
            (req.branch_id, student["id"], student["name"], parse_iso_date(req.date), req.status, req.lecture),
        )
    return {"status": "success", "student_id": student["id"]}


@app.get("/api/attendance/{branch_id}/{date}")
def attendance_for_date(branch_id: int, date: str, user: CurrentInstitute = Depends(get_current_institute)):
    require_module_access(user, "attendance")
    verify_branch_ownership(branch_id, user.id)
    with db_cursor() as (_, cur):
        cur.execute("SELECT student_id,student_name,status FROM attendance WHERE branch_id=%s AND date=%s", (branch_id, parse_iso_date(date)))
        return {str(r["student_id"]): r["status"] for r in cur.fetchall()}


@app.get("/api/attendance/student/{student_id}/history")
def attendance_history(student_id: int, branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    require_module_access(user, "attendance")
    verify_branch_ownership(branch_id, user.id)
    with db_cursor() as (_, cur):
        cur.execute("SELECT id,date,status,lecture FROM attendance WHERE student_id=%s AND branch_id=%s ORDER BY date DESC", (student_id, branch_id))
        rows = cur.fetchall()
        cur.execute("SELECT id,name,batch,roll_number FROM students WHERE id=%s AND branch_id=%s", (student_id, branch_id))
        student = cur.fetchone()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    total = len(rows)
    present = sum(1 for r in rows if r["status"] == "Present")
    return {"student": student, "summary": {"present": present, "absent": total - present, "total": total, "percentage": round(present * 100 / total, 1) if total else 0}, "history": rows}


# ---------------------------------------------------------------------------
# Timetable solver
# ---------------------------------------------------------------------------

class TimetableGenerateRequest(BaseModel):
    branch_id: int
    batch_name: str
    teachers_config: list[dict[str, Any]]
    timings: list[Any]
    days: list[str] = Field(default_factory=lambda: ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"])
    unavailable_days: list[str] = Field(default_factory=list)


class TimetableSlotUpdate(BaseModel):
    day: Optional[str] = None
    lecture_number: Optional[int] = None
    time_slot: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    subject: Optional[str] = None
    teacher: Optional[str] = None
    room: Optional[str] = None


@app.get("/api/timetable/slots/{branch_id}")
def get_timetable_slots(branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    require_module_access(user, "timetable")
    verify_branch_ownership(branch_id, user.id)
    with db_cursor() as (_, cur):
        cur.execute("SELECT * FROM timetables_slots WHERE branch_id=%s ORDER BY batch_name, day, lecture_number, start_time", (branch_id,))
        return cur.fetchall()


@app.get("/api/timetable/generations/{branch_id}")
def timetable_generations(branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    require_module_access(user, "timetable")
    verify_branch_ownership(branch_id, user.id)
    with db_cursor() as (_, cur):
        cur.execute("SELECT DISTINCT ON(batch_name) id,batch_name,config,created_at FROM timetable_generations WHERE branch_id=%s ORDER BY batch_name,created_at DESC", (branch_id,))
        return cur.fetchall()


def solve_timetable(cur, req: TimetableGenerateRequest, user: CurrentInstitute) -> tuple[list[dict], list[dict], dict]:
    batch = req.batch_name.strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Batch name is required")
    if not req.teachers_config:
        raise HTTPException(status_code=400, detail="Add at least one teacher and subject")
    timings = [timing_from_input(v, i) for i, v in enumerate(req.timings)]
    numbers = [t["lecture_number"] for t in timings]
    if len(set(numbers)) != len(numbers):
        raise HTTPException(status_code=400, detail="Lecture numbers must be unique")
    for left_index, left in enumerate(timings):
        left_start, left_end = parse_clock(left["start_time"]), parse_clock(left["end_time"])
        for right in timings[left_index + 1:]:
            right_start, right_end = parse_clock(right["start_time"]), parse_clock(right["end_time"])
            if left_start < right_end and right_start < left_end:
                raise HTTPException(status_code=400, detail="Lecture timings cannot overlap")
    days = [d for d in req.days if d in {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"}]
    days = [d for d in days if d not in set(req.unavailable_days)]
    if not days:
        raise HTTPException(status_code=400, detail="At least one available day is required")
    cur.execute("SELECT room_no FROM classrooms WHERE branch_id=%s AND room_no IS NOT NULL ORDER BY id", (req.branch_id,))
    rooms = [r["room_no"] for r in cur.fetchall()]
    # The batch is regenerated atomically: old rows are removed inside this
    # transaction, while rows from other batches remain hard constraints.
    cur.execute("DELETE FROM timetables_slots WHERE branch_id=%s AND LOWER(batch_name)=LOWER(%s)", (req.branch_id, batch))
    cur.execute("SELECT day,time_slot,teacher,room FROM timetables_slots WHERE branch_id=%s", (req.branch_id,))
    existing = cur.fetchall()
    batch_busy: set[tuple] = set()
    teacher_busy: set[tuple] = {(r["day"], r["time_slot"], r["teacher"]) for r in existing}
    room_busy: set[tuple] = {(r["day"], r["time_slot"], r["room"]) for r in existing if r["room"]}
    units = []
    for config in req.teachers_config:
        name = str(config.get("name", "")).strip()
        subject = str(config.get("subject", "")).strip()
        target = int(config.get("lectures_per_week", 0))
        if not name or not subject or target < 1:
            raise HTTPException(status_code=400, detail="Every teacher row needs a name, subject and lectures per week")
        unavailable = set(config.get("unavailable_days") or []) | set(req.unavailable_days)
        candidates = [(d, t) for d in days if d not in unavailable for t in timings]
        if target > len(candidates):
            raise HTTPException(status_code=409, detail=f"{name} needs {target} lectures but only {len(candidates)} valid day/timing combinations remain")
        for occurrence in range(1, target + 1):
            units.append({"name": name, "subject": subject, "candidates": candidates, "occurrence": occurrence})
    # Most constrained teachers go first. Backtracking means the solver never
    # silently schedules fewer lectures than requested.
    units.sort(key=lambda u: len(u["candidates"]))
    chosen: list[dict] = []

    def backtrack(index: int) -> bool:
        if index == len(units):
            return True
        unit = units[index]
        for day, timing in unit["candidates"]:
            key = (day, timing["label"])
            if key in batch_busy or (day, timing["label"], unit["name"]) in teacher_busy:
                continue
            room = None
            if rooms:
                for candidate in rooms:
                    if (day, timing["label"], candidate) not in room_busy:
                        room = candidate
                        break
                if room is None:
                    continue
            batch_busy.add(key)
            teacher_key = (day, timing["label"], unit["name"])
            teacher_busy.add(teacher_key)
            if room:
                room_busy.add((day, timing["label"], room))
            chosen.append({"day": day, "timing": timing, "lecture_number": timing["lecture_number"], "time_slot": timing["label"], "subject": unit["subject"], "teacher": unit["name"], "room": room})
            if backtrack(index + 1):
                return True
            chosen.pop()
            batch_busy.remove(key)
            teacher_busy.remove(teacher_key)
            if room:
                room_busy.remove((day, timing["label"], room))
        return False

    if not backtrack(0):
        raise HTTPException(status_code=409, detail="No timetable satisfies every lecture count, unavailable day, teacher, batch and classroom constraint. Nothing was changed.")
    config = {"batch_name": batch, "teachers_config": req.teachers_config, "timings": timings, "days": days, "unavailable_days": req.unavailable_days}
    return chosen, timings, config


@app.post("/api/timetable/generate")
def generate_timetable(req: TimetableGenerateRequest, user: CurrentInstitute = Depends(require_write_access)):
    require_module_access(user, "timetable", write=True)
    verify_branch_ownership(req.branch_id, user.id)
    # Transaction-scoped advisory lock serialises two regenerate clicks for the
    # same branch/batch. Thus generations can never be stacked or interleaved.
    conn = get_conn()
    try:
        with conn.transaction():
            with conn.cursor() as cur:
                # Use a branch-wide lock so two different batches cannot race
                # while checking shared teachers/classrooms either.
                cur.execute("SELECT pg_advisory_xact_lock(%s, %s)", (0, req.branch_id))
                chosen, _, config = solve_timetable(cur, req, user)
                cur.execute("INSERT INTO timetable_generations(branch_id,batch_name,config,created_by) VALUES(%s,%s,%s::jsonb,%s) RETURNING id", (req.branch_id, req.batch_name.strip(), json.dumps(config), user.id))
                generation_id = cur.fetchone()["id"]
                for item in chosen:
                    t = item["timing"]
                    cur.execute(
                        """INSERT INTO timetables_slots(branch_id,generation_id,batch_name,day,time_slot,lecture_number,start_time,end_time,subject,teacher,room)
                           VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (req.branch_id, generation_id, req.batch_name.strip(), item["day"], t["label"], t["lecture_number"], t["start_time"], t["end_time"], item["subject"], item["teacher"], item["room"]),
                    )
                cur.execute("DELETE FROM timetable_generations WHERE branch_id=%s AND LOWER(batch_name)=LOWER(%s) AND id<>%s", (req.branch_id, req.batch_name.strip(), generation_id))
        return {"status": "success", "generation_id": generation_id, "slots": chosen, "warnings": []}
    except HTTPException:
        conn.rollback()
        raise
    finally:
        conn.close()


@app.patch("/api/timetable/slots/{slot_id}")
def edit_timetable_slot(slot_id: int, req: TimetableSlotUpdate, user: CurrentInstitute = Depends(require_write_access)):
    require_module_access(user, "timetable", write=True)
    with db_cursor() as (_, cur):
        cur.execute("SELECT s.* FROM timetables_slots s JOIN branches b ON b.id=s.branch_id WHERE s.id=%s AND b.institute_id=%s", (slot_id, user.id))
        old = cur.fetchone()
        if not old:
            raise HTTPException(status_code=404, detail="Timetable slot not found")
        data = model_data(req, exclude_none=True)
        if "start_time" in data or "end_time" in data:
            start = parse_clock(data.get("start_time", old["start_time"].strftime("%H:%M")))
            end = parse_clock(data.get("end_time", old["end_time"].strftime("%H:%M")))
            if start >= end:
                raise HTTPException(status_code=400, detail="End time must be after start time")
            data["start_time"], data["end_time"] = start, end
            data["time_slot"] = f"{start.strftime('%I:%M %p')} – {end.strftime('%I:%M %p')}".replace(" 0", " ")
        if not data:
            raise HTTPException(status_code=400, detail="No timetable fields supplied")
        # A manual edit is also constraint-aware: it may not move a lecture to
        # a holiday, an unavailable teacher day, or a timing/teacher outside
        # the generation configuration.
        cur.execute("SELECT config FROM timetable_generations WHERE id=%s AND branch_id=%s", (old["generation_id"], old["branch_id"]))
        generation = cur.fetchone()
        if generation:
            config = generation["config"] or {}
            selected_day = data.get("day", old["day"])
            selected_teacher = data.get("teacher", old["teacher"])
            if selected_day not in (config.get("days") or []):
                raise HTTPException(status_code=409, detail="That day is not available in this timetable")
            configured_teacher = next((item for item in config.get("teachers_config", []) if item.get("name") == selected_teacher), None)
            if configured_teacher is None:
                raise HTTPException(status_code=409, detail="That teacher is not part of this timetable configuration")
            if selected_day in set(configured_teacher.get("unavailable_days") or []) | set(config.get("unavailable_days") or []):
                raise HTTPException(status_code=409, detail="That teacher is unavailable on the selected day")
            if "lecture_number" in data and data["lecture_number"] not in {int(t.get("lecture_number")) for t in config.get("timings", [])}:
                raise HTTPException(status_code=409, detail="Use one of the configured lecture numbers")
        # Enforce batch, teacher and room collision checks before an edit,
        # using interval overlap rather than merely matching a display label.
        day = data.get("day", old["day"])
        slot = data.get("time_slot", old["time_slot"])
        teacher = data.get("teacher", old["teacher"])
        room = data.get("room", old["room"])
        start_for_check = data.get("start_time", old["start_time"])
        end_for_check = data.get("end_time", old["end_time"])
        cur.execute("SELECT 1 FROM timetables_slots WHERE branch_id=%s AND id<>%s AND batch_name=%s AND day=%s AND start_time < %s AND end_time > %s", (old["branch_id"], slot_id, old["batch_name"], day, end_for_check, start_for_check))
        if cur.fetchone():
            raise HTTPException(status_code=409, detail="That batch already has a lecture at this time")
        cur.execute("SELECT 1 FROM timetables_slots WHERE branch_id=%s AND id<>%s AND day=%s AND start_time < %s AND end_time > %s AND teacher=%s", (old["branch_id"], slot_id, day, end_for_check, start_for_check, teacher))
        if cur.fetchone():
            raise HTTPException(status_code=409, detail="That teacher is already assigned at this time")
        if room:
            cur.execute("SELECT 1 FROM timetables_slots WHERE branch_id=%s AND id<>%s AND day=%s AND start_time < %s AND end_time > %s AND room=%s", (old["branch_id"], slot_id, day, end_for_check, start_for_check, room))
            if cur.fetchone():
                raise HTTPException(status_code=409, detail="That classroom is already occupied at this time")
        parts, vals = [], []
        for k, v in data.items():
            parts.append(f"{k}=%s")
            vals.append(v)
        vals.append(slot_id)
        cur.execute(f"UPDATE timetables_slots SET {', '.join(parts)} WHERE id=%s", vals)
    return {"status": "updated", "id": slot_id}


# ---------------------------------------------------------------------------
# Dashboard analytics
# ---------------------------------------------------------------------------

@app.get("/api/dashboard/{branch_id}")
def dashboard(branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    require_module_access(user, "dashboard")
    verify_branch_ownership(branch_id, user.id)
    now = datetime.now(timezone.utc)
    week_start = (now.date() - timedelta(days=6))
    with db_cursor() as (_, cur):
        cur.execute("SELECT COUNT(*) AS n FROM students WHERE branch_id=%s AND COALESCE(status,'Active') NOT IN ('Inactive','Archived')", (branch_id,)); students = cur.fetchone()["n"]
        cur.execute("SELECT COUNT(*) AS n FROM teachers WHERE branch_id=%s", (branch_id,)); teachers = cur.fetchone()["n"]
        cur.execute("SELECT COUNT(*) AS n FROM classrooms WHERE branch_id=%s", (branch_id,)); classrooms = cur.fetchone()["n"]
        cur.execute("SELECT COALESCE(SUM(amount_inr),0) AS total, COUNT(*) AS count FROM fees WHERE branch_id=%s AND LOWER(status) NOT IN ('paid','settled')", (branch_id,)); fee_row = cur.fetchone()
        cur.execute("SELECT batch, COUNT(*) AS enrolled FROM students WHERE branch_id=%s GROUP BY batch ORDER BY batch", (branch_id,)); batch_rows = cur.fetchall()
        cur.execute("""SELECT s.batch,a.status,COUNT(*) AS n FROM attendance a JOIN students s ON s.id=a.student_id
                       WHERE a.branch_id=%s AND a.date BETWEEN %s AND %s GROUP BY s.batch,a.status""", (branch_id, week_start, now.date())); attendance_rows = cur.fetchall()
        cur.execute("SELECT * FROM timetables_slots WHERE branch_id=%s AND day=%s", (branch_id, now.strftime("%A"),)); today_slots = cur.fetchall()
        cur.execute("SELECT student_name,amount_inr,due_date,status FROM fees WHERE branch_id=%s AND LOWER(status) NOT IN ('paid','settled') ORDER BY due_date NULLS LAST LIMIT 8", (branch_id,)); pending_fees = cur.fetchall()
    attendance_by_batch = []
    for row in batch_rows:
        marks = [r for r in attendance_rows if r["batch"] == row["batch"]]
        present = sum(r["n"] for r in marks if r["status"] == "Present")
        absent = sum(r["n"] for r in marks if r["status"] == "Absent")
        total = present + absent
        attendance_by_batch.append({"batch": row["batch"], "enrolled": row["enrolled"], "present": present, "absent": absent, "marked": total, "percentage": round(present * 100 / total, 1) if total else 0})
    ongoing = []
    current_time = now.time().replace(tzinfo=None)
    for slot in today_slots:
        start = slot["start_time"]
        end = slot["end_time"]
        if start <= current_time <= end:
            ongoing.append({"id": slot["id"], "batch_name": slot["batch_name"], "subject": slot["subject"], "teacher": slot["teacher"], "room": slot["room"], "time_slot": slot["time_slot"], "lecture_number": slot["lecture_number"]})
    ongoing_by_batch = [{"batch": row["batch"], "lectures": [lecture for lecture in ongoing if lecture["batch_name"] == row["batch"]]} for row in batch_rows]
    return {"date": now.date(), "viewed_at": now.isoformat(), "summary": {"students": students, "teachers": teachers, "classrooms": classrooms, "fees_pending": float(fee_row["total"] or 0), "fee_records_pending": fee_row["count"]}, "attendance_this_week": attendance_by_batch, "ongoing_lectures": ongoing, "ongoing_by_batch": ongoing_by_batch, "pending_fees": pending_fees}


# ---------------------------------------------------------------------------
# XLSX and PDF exports
# ---------------------------------------------------------------------------


def export_rows(module: str, branch_id: int, user: CurrentInstitute) -> tuple[list[str], list[list[Any]]]:
    if module in RECORD_MODULES:
        require_module_access(user, module)
        with db_cursor() as (_, cur):
            cur.execute(f"SELECT * FROM {TABLES[module]} WHERE branch_id=%s ORDER BY id", (branch_id,))
            rows = cur.fetchall()
        if not rows:
            return ["ID"], []
        headers = list(rows[0].keys())
        return headers, [[str(r.get(h) if r.get(h) is not None else "") for h in headers] for r in rows]
    if module == "timetable":
        require_module_access(user, module)
        with db_cursor() as (_, cur):
            cur.execute("SELECT batch_name,day,lecture_number,time_slot,subject,teacher,room FROM timetables_slots WHERE branch_id=%s ORDER BY batch_name,day,lecture_number", (branch_id,))
            rows = cur.fetchall()
        headers = ["batch_name", "day", "lecture_number", "time_slot", "subject", "teacher", "room"]
        return headers, [[r.get(h) or "" for h in headers] for r in rows]
    if module == "seating":
        require_module_access(user, module)
        return ["Status"], [["Seating matrix uses the current student/classroom data"]]
    raise HTTPException(status_code=400, detail="Invalid export module")


@app.get("/api/export/{module}/{branch_id}.xlsx")
def export_xlsx(module: str, branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    verify_branch_ownership(branch_id, user.id)
    headers, rows = export_rows(module, branch_id, user)
    wb = Workbook(); ws = wb.active; ws.title = module[:31].upper()
    ws.append([h.replace("_", " ").title() for h in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="8A641C")
    for row in rows: ws.append(row)
    ws.freeze_panes = "A2"
    for column in ws.columns:
        letter = column[0].column_letter
        ws.column_dimensions[letter].width = min(max(max(len(str(c.value or "")) for c in column) + 2, 12), 32)
    output = io.BytesIO(); wb.save(output); output.seek(0)
    filename = f"algorithmic-{module}-{branch_id}.xlsx"
    return Response(output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/api/export/{module}/{branch_id}.pdf")
def export_pdf(module: str, branch_id: int, user: CurrentInstitute = Depends(get_current_institute)):
    verify_branch_ownership(branch_id, user.id)
    headers, rows = export_rows(module, branch_id, user)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    data = [[Paragraph(str(h).replace("_", " ").title(), styles["Normal"]) for h in headers]] + [[Paragraph(str(c), styles["Normal"]) for c in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8A641C")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("FONTSIZE", (0, 0), (-1, -1), 7), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
    doc.build([Paragraph(f"ALGORITHMIC — {module.upper()}", styles["Title"]), table])
    return Response(output.getvalue(), media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="algorithmic-{module}-{branch_id}.pdf"'})


# ---------------------------------------------------------------------------
# Frontend
# ---------------------------------------------------------------------------

@app.get("/", response_class=FileResponse)
def root():
    return FileResponse(BASE_DIR / "index.html", media_type="text/html")


@app.get("/uploads/{filename}")
def uploaded_file(filename: str):
    safe = Path(filename).name
    path = UPLOAD_DIR / safe
    if not path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path)
